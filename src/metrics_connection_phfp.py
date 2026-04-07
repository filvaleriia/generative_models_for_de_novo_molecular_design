import pandas as pd
import argparse
from sklearn.preprocessing import MinMaxScaler
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.path_utils import pharm_results_dir



def connect_mean_value(type_cluster, type_phfp, generators_name_list, receptor, threshold=1, data_folder=''):
    """
    Combines mean values from multiple generators into a single DataFrame.

    Parameters:
    - receptor: Target receptor
    - type_phfp: Type of pharmacophore fingerprint
    - type_cluster: Cluster type
    - generators_name_list: List of generator names
    - threshold: Data threshold to be analyzed

    Returns:
    - A combined DataFrame with mean values
    """

    # Define path to data
    link = pharm_results_dir(data_folder, receptor, type_phfp, type_cluster)

    # List to store paths of mean value files
    link_mean = []
    for generator in generators_name_list:
        link_mean.append(
            link / generator / f"threshold_{threshold}" / f"{generator}_mean_{type_phfp}_{type_cluster}_threshold_{threshold}.csv"
        )

    # Load data and merge into a single DataFrame
    df_list = [pd.read_csv(f) for f in link_mean]
    df = pd.concat(df_list, ignore_index=True)

    # Save results to CSV files
    output_path = link / f"mean_{type_phfp}_{type_cluster}_threshold_{threshold}.csv"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_path, index=False)
    return df



def connect_mean_value_normalized(type_cluster, type_phfp, generators_name_list, receptor, threshold, data_folder=''):
    """
    Loads and normalizes mean values using Min-Max scaling.

    Parameters:
    - receptor: Target receptor
    - type_phfp: Type of pharmacophore fingerprint
    - type_cluster: Cluster type
    - generators_name_list: List of generator names
    - threshold: Data threshold to be analyzed

    Returns:
    - Normalized DataFrame
    """

    # Define path to data
    link = pharm_results_dir(data_folder, receptor, type_phfp, type_cluster)

    # List to store paths of mean value files
    link_mean = [
        link / generator / f"threshold_{threshold}" / f"{generator}_mean_{type_phfp}_{type_cluster}_threshold_{threshold}.csv"
        for generator in generators_name_list
    ]

    # Load data
    df_list = [pd.read_csv(f) for f in link_mean]
    df = pd.concat(df_list, ignore_index=True)

    # Normalize using Min-Max scaling
    scaler = MinMaxScaler()
    numeric_columns = df.select_dtypes(include=['number']).columns
    df[numeric_columns] = scaler.fit_transform(df[numeric_columns])

    # Save normalized results
    output_path = link / f"mean_{type_phfp}_{type_cluster}_threshold_{threshold}_norm_min_max.csv"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_path, index=False)
    return df



def prettify_generator_name(name: str) -> str:
    if not isinstance(name, str):
        return name

    name = name.replace("_mean", "")

    if name.startswith("DrugEx_GT_epsilon_"):
        eps = name.split("DrugEx_GT_epsilon_")[-1]
        return f"DrugEx GT\nepsilon {eps}"

    if name.startswith("DrugEx_RNN_epsilon_"):
        eps = name.split("DrugEx_RNN_epsilon_")[-1]
        return f"DrugEx RNN\nepsilon {eps}"

    if name.startswith("GB_GA_mut_r_"):
        rate = name.split("GB_GA_mut_r_")[-1]
        return f"GB_GA\nmut.r. {rate}"

    if name.startswith("GB_GA_log_p_mut_r_"):
        rate = name.split("GB_GA_log_p_mut_r_")[-1]
        return f"GB_GA log_p\nmut.r. {rate}"

    if name.startswith("addcarbon"):
        return "AddCarbon"
    
    if name.startswith("enamine"):
        return "Enamine"

    return name.replace("_", " ")



def export_article_excel_table(
    df: pd.DataFrame,
    out_xlsx: str,
    receptor: str,
    type_phfp: str,
    threshold: str | int | float,
    type_cluster: str | None = None,
):
    """
    Creates an Excel table suitable for a paper for pharmacophore fingerprint results.

    Formatting rules:
      - RS, SED: 2 decimals
      - ASER: displayed in units of ×10^-2 (ASER * 100), 3 decimals
      - Column header for ASER is "ASER · 10^-2"
      - Maximum value in each metric column is bold within each block (Split dis/sim)
      - Header styling, borders, alignment
      - Supports Split column if present; otherwise uses type_cluster
    """

    out_path = Path(out_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    df = df.copy()

    if type_phfp == 'rdkit':
        type_phfp_str = 'RDKit'

    if "PHFP" not in df.columns:
        df["PHFP"] = type_phfp_str

    if "Threshold" not in df.columns:
        df["Threshold"] = threshold

    if "Split" not in df.columns:
        df["Split"] = type_cluster if type_cluster is not None else ""

    if "name" in df.columns and "Name" not in df.columns:
        df = df.rename(columns={"name": "Name"})

    df["Name"] = df["Name"].apply(prettify_generator_name)
    df["PHFP"] = type_phfp_str

    wanted_cols = ["Name", "RS", "SED", "ASER", "PHFP", "Threshold", "Split"]
    existing_cols = [c for c in wanted_cols if c in df.columns]
    df = df[existing_cols]

    metric_cols = [c for c in ["RS", "SED", "ASER"] if c in df.columns]

    for c in ["RS", "SED"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(2)

    if "ASER" in df.columns:
        df["ASER"] = pd.to_numeric(df["ASER"], errors="coerce") * 100.0
        df["ASER"] = df["ASER"].round(3)

    header_labels = {
        "Name": "Name",
        "RS": "RS",
        "SED": "SED",
        "ASER": "ASER · 10⁻²",
        "PHFP": "PHFP",
        "Threshold": "Threshold",
        "Split": "Split",
    }

    wb = Workbook()
    ws = wb.active
    sheet_name = f"{receptor}_{type_phfp}"
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="EDEDED")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_block(start_row: int, block_title: str, block_df: pd.DataFrame) -> int:
        r = start_row
        ncols = len(existing_cols)

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1, value=block_title)
        cell.font = title_font
        cell.alignment = left
        r += 1

        for j, col in enumerate(existing_cols, start=1):
            label = header_labels.get(col, col)
            c = ws.cell(row=r, column=j, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        r += 1

        maxima = {}
        for mc in metric_cols:
            if mc in block_df.columns:
                maxima[mc] = block_df[mc].max(skipna=True)

        for _, row in block_df.iterrows():
            for j, col in enumerate(existing_cols, start=1):
                val = row[col]
                cell = ws.cell(row=r, column=j, value=val)
                cell.alignment = center
                cell.border = border

                if col in metric_cols:

                    cell.number_format = "0.00"

                    try:
                        if pd.notna(val) and pd.notna(maxima.get(col)) and float(val) == float(maxima[col]):
                            cell.font = bold_font
                    except Exception:
                        pass
            r += 1

        return r + 1

    split_order = ["dis", "sim"]
    unique_splits = df["Split"].astype(str).unique().tolist()
    splits = [s for s in split_order if s in unique_splits]
    splits += [s for s in unique_splits if s not in splits]

    current_row = 1
    caption = f"Receptor: {receptor} | PHFP: {type_phfp_str} | Threshold: {threshold}"
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(existing_cols))
    cap_cell = ws.cell(row=current_row, column=1, value=caption)
    cap_cell.font = Font(bold=True, size=13)
    cap_cell.alignment = left
    current_row += 2

    for sp in splits:
        block_df = df[df["Split"].astype(str) == str(sp)].copy()
        block_title = f"Split: {sp}" if sp else "Results"
        current_row = write_block(current_row, block_title, block_df)

    col_widths = {
        "Name": 22,
        "RS": 10,
        "SED": 10,
        "ASER": 12,
        "PHFP": 12,
        "Threshold": 12,
        "Split": 8,
    }
    for j, col in enumerate(existing_cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = col_widths.get(col, 12)

    ws.freeze_panes = "A4"

    wb.save(out_xlsx)
    print(f"[OK] Excel saved to: {out_xlsx}")



def resolve_threshold(type_cluster, threshold=None, threshold_sim=None, threshold_dis=None):
    """Return threshold according to split, allowing dis/sim thresholds to differ."""
    if threshold is not None:
        return threshold

    if type_cluster == "sim" and threshold_sim is not None:
        return threshold_sim

    if type_cluster == "dis" and threshold_dis is not None:
        return threshold_dis

    raise ValueError(
        "Threshold is not defined. Use --threshold or provide --threshold_sim / --threshold_dis for the given split."
    )



def main():
    parser = argparse.ArgumentParser(description='Compute and visualize pharmacophore fingerprint metrics.')
    parser.add_argument('--type_cluster', type=str, required=True, help='Type of clustering (dis/sim)')
    parser.add_argument('--type_phfp', type=str, required=True, help='Type of pharmacophore fingerprint')
    parser.add_argument('--generator_list', nargs='+', required=True, help='Generator name')
    parser.add_argument('--receptor', type=str, required=True, help='Receptor name')
    parser.add_argument('--threshold', type=str, required=False, help='Single threshold for backward compatibility')
    parser.add_argument('--threshold_sim', type=str, required=False, help='Threshold for sim split')
    parser.add_argument('--threshold_dis', type=str, required=False, help='Threshold for dis split')
    parser.add_argument('--data_folder', type=str, required=False, default='', help='Data dir')

    args = parser.parse_args()

    threshold = resolve_threshold(
        type_cluster=args.type_cluster,
        threshold=args.threshold,
        threshold_sim=args.threshold_sim,
        threshold_dis=args.threshold_dis,
    )

    df_raw = connect_mean_value(
        args.type_cluster,
        args.type_phfp,
        args.generator_list,
        args.receptor,
        threshold,
        args.data_folder,
    )
    connect_mean_value_normalized(
        args.type_cluster,
        args.type_phfp,
        args.generator_list,
        args.receptor,
        threshold,
        args.data_folder,
    )

    out_xlsx = str(
        pharm_results_dir(args.data_folder, args.receptor, args.type_phfp, args.type_cluster)
        / f"paper_table_{args.receptor}_{args.type_phfp}_{args.type_cluster}_threshold_{threshold}.xlsx"
    )

    export_article_excel_table(
        df=df_raw,
        out_xlsx=out_xlsx,
        receptor=args.receptor,
        type_phfp=args.type_phfp,
        threshold=threshold,
        type_cluster=args.type_cluster,
    )


if __name__ == "__main__":
    main()
